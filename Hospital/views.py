from django.db.models import Max
from django.shortcuts import render,redirect
from app.models import Patient,Appointment,Doctor,Room,Patients_Chart,Appointment_Chart,Patient_Payment
from django.http import HttpResponse,HttpResponseRedirect,FileResponse
import io,csv
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,login
from django.shortcuts import render
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime
from django.db.models import Sum



def P_Venue_Pdf(request):
    buf = io.BytesIO()
    c = canvas.Canvas(buf,pagesize=letter,bottomup=0)
    textob = c.beginText()
    textob.setTextOrigin(inch, inch)
    textob.setFont("Helvetica", 14)

    patient = Patient.objects.all()
    lines = []
    for n in patient:
        lines.append(n.email)
        lines.append(n.patient_name)
        lines.append(n.gender)
        #lines.append(n.phone)
        lines.append("=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=")
    for line in lines:
        textob.textLine(line)
    c.drawText(textob)
    c.showPage()
    c.save()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='venue.pdf')


def D_Venue_Pdf(request):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    textob = c.beginText()
    textob.setTextOrigin(inch, inch)
    textob.setFont("Helvetica", 14)

    doctor = Doctor.objects.all()
    lines = []
    for n in doctor:
        lines.append(n.doctor_name)
        lines.append(n.dob)
        lines.append(n.specialization)
        lines.append(n.experience)
        lines.append(n.email)
        # lines.append(n.phone)
        lines.append("=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=")
    for line in lines:
        textob.textLine(line)
    c.drawText(textob)
    c.showPage()
    c.save()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='venue.pdf')


def A_Venue_Pdf(request):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    textob = c.beginText()
    textob.setTextOrigin(inch, inch)
    textob.setFont("Helvetica", 14)

    appointment = Appointment.objects.all()
    lines = []
    for n in appointment:
        lines.append(n.department)
        lines.append(n.doctor_name)
        lines.append(n.appointment_data)
        lines.append(n.time_slot)
        lines.append(n.problem)
        # lines.append(n.phone)
        lines.append("=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=")
    for line in lines:
        textob.textLine(line)
    c.drawText(textob)
    c.showPage()
    c.save()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='venue.pdf')


def R_Venue_Pdf(request):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    textob = c.beginText()
    textob.setTextOrigin(inch, inch)
    textob.setFont("Helvetica", 14)

    room = Room.objects.all()
    lines = []
    for n in room:
        lines.append(n.room_type)
        lines.append(n.patient_name)
        lines.append(n.allotment_date)
        lines.append(n.discharge_date)
        lines.append(n.doctor_name)
        # lines.append(n.phone)
        lines.append("=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=")
    for line in lines:
        textob.textLine(line)
    c.drawText(textob)
    c.showPage()
    c.save()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='venue.pdf')

def PP_Venue_Pdf(request):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    textob = c.beginText()
    textob.setTextOrigin(inch, inch)
    textob.setFont("Helvetica", 14)

    room = Patient_Payment.objects.all()
    lines = []
    for n in room:
        lines.append(n.invoice_date)
        lines.append(n.patient_name)
        lines.append(n.department)
        lines.append(n.doctor_name)
        lines.append(n.admission_date)
        lines.append(n.discharge_date)
        lines.append(n.service_name)
        lines.append(n.advance_paid)
        lines.append(n.payment_type)
        lines.append(n.card_no)
        # lines.append(n.phone)
        lines.append("=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=")
    for line in lines:
        textob.textLine(line)
    c.drawText(textob)
    c.showPage()
    c.save()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='venue.pdf')


def P_Venue_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=venue.csv'

    writer = csv.writer(response)
    patient = Patient.objects.all()
    writer.writerow(['Patient Id','Patient Name','Patient Email','Patient Gender','Patient DOB','Patient Age','Phone','Address'])
    lines = []
    for n in patient:
        writer.writerow([n.patient_id,n.patient_name,n.email,n.gender,n.dob,n.age,n.phone,n.address])

    return response

def D_Venue_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=venue.csv'

    writer = csv.writer(response)
    doctor = Doctor.objects.all()
    writer.writerow(['Doctor Id','Doctor Name','Doctor Email','Doctor Gender','Doctor DOB','Doctor Age','Specialization','Experience','Phone','Address'])
    for n in doctor:
        writer.writerow([n.doctor_id,n.doctor_name,n.email,n.gender,n.dob,n.age,n.specialization,n.experience,n.phone,n.address])

    return response

def A_Venue_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=venue.csv'

    writer = csv.writer(response)
    appointment = Appointment.objects.all()
    writer.writerow(['Patient Id','Appointment Id','Department','Doctor Name','Appointment Date','Time Slot','Token No','Problem'])
    lines = []
    for n in appointment:
        writer.writerow([n.patient_id,n.appointment_id,n.department,n.doctor_name,n.appointment_data,n.time_slot,n.token_no,n.problem])

    return response

def R_Venue_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=venue.csv'

    writer = csv.writer(response)
    room = Room.objects.all()
    writer.writerow(['Room No','Room Type','Patient Name','Allotment Date','Discharge Date','Doctor Name'])
    lines = []
    for n in room:
        writer.writerow([n.room_no,n.room_type,n.patient_name,n.allotment_date,n.discharge_date,n.doctor_name])

    return response

def PP_Venue_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=venue.csv'

    writer = csv.writer(response)
    room = Patient_Payment.objects.all()
    writer.writerow(['Invoice id','Invoice date','Patient id','Patient name','Department','Doctor name','Admission date','Discharge date','Service name','Cost of treatment','Discount','Advance paid','Payment type','Card no'])
    lines = []
    for n in room:
        writer.writerow([n.invoice_id,n.invoice_date,n.patient_id,n.patient_name,n.department,n.doctor_name,n.admission_date,n.discharge_date,n.service_name,n.cost_of_treatment,n.discount,n.advance_paid,n.payment_type,n.card_no])

    return response


def BASE(request):
    return render(request,'base.html')


def ADD_PATIENT(request):
    if request.session.has_key('id'):
        if request.method == "POST":
            id = 1001 if Patient.objects.count() == 0 else Patient.objects.aggregate(max=Max('patient_id'))["max"] + 1
            patient_name = request.POST.get('patient_name')
            dob = request.POST.get('dob')
            age = request.POST.get('age')
            phone = request.POST.get('phone')
            email = request.POST.get('email')
            gender = request.POST.get('gender')
            address = request.POST.get('address')

            patient=Patient(
                patient_id = id,
                patient_name = patient_name,
                dob = dob,
                age = age,
                phone = phone,
                email = email,
                gender = gender,
                address = address
            )
            patient.save()
        return render(request,'patients/add_patient.html')
    else:
        return redirect('')


def ABOUT_PATIENT(request):
    if request.session.has_key('id'):
        data = Patient.objects.all()
        data1 = Patient_Payment.objects.all()
        if request.method == "POST":
            patient_name = request.POST.get('patient_name')
            dob = request.POST.get('dob')
            age = request.POST.get('age')
            phone = request.POST.get('phone')
            email = request.POST.get('email')
            gender = request.POST.get('gender')
            address = request.POST.get('address')
            invoice_id = request.POST.get('invoice_id')
            invoice_date = datetime.now()
            patient_name = request.POST.get('patient_name')
            doctor_name = request.POST.get('doctor_name')
            service_name = request.POST.get('service_name')
            cost_of_treatment = request.POST.get('cost_of_treatment')
            discount = request.POST.get('discount')

             
            data = Patient(patient_name,dob,age,phone,email,gender,address)
            data1 = Patient_Payment(invoice_id,invoice_date,patient_name,doctor_name,service_name,cost_of_treatment,discount)
            
            data1.save()
            data.save()
          
        return render(request,'patients/about_patient.html',{'details':data, 'details1': data1})
    else:
        return redirect('')


def ALL_PATIENT(request):
    if request.session.has_key('id'):
        data = Patient.objects.all()
        if request.method == "POST":
            patient_id = request.POST.get('patient_id')
            patient_name = request.POST.get('patient_name')
            dob = request.POST.get('dob')
            age = request.POST.get('age')
            phone = request.POST.get('phone')
            email = request.POST.get('email')
            gender = request.POST.get('gender')
            address = request.POST.get('address')

            data = Patient(patient_id,patient_name, dob, age, phone, email, gender, address)
            data.save()
        return render(request, 'patients/all_patient.html', {"details": data})
    else:
        return redirect('')


def EDIT_PATIENT(request,patient_id):
    patient = Patient.objects.get(patient_id=patient_id)

    return render(request,'patients/edit_patient.html',{'Patient':patient})

def Update_Patient(request,patient_id):
    patient_name = request.POST.get('patient_name')
    dob = request.POST.get('dob')
    age = request.POST.get('age')
    phone = request.POST.get('phone')
    email = request.POST.get('email')
    gender = request.POST.get('gender')
    address = request.POST.get('address')
    my_user = Patient.objects.get(patient_id=patient_id)
    my_user.patient_name = patient_name
    my_user.dob = dob
    my_user.age = age
    my_user.phone = phone
    my_user.email = email
    my_user.gender = gender
    my_user.address = address

    my_user.save()
    return redirect('all_patient')

def ADD_DOCTOR(request):
    if request.session.has_key('id'):
        if request.method == "POST":
            doctor_id = 1001 if Doctor.objects.count() == 0 else Doctor.objects.aggregate(max=Max('doctor_id'))["max"] + 1
            doctor_name = request.POST.get('doctor_name')
            dob = request.POST.get('dob')
            specialization = request.POST.get('specialization')
            experience = request.POST.get('experience')
            age = request.POST.get('age')
            phone = request.POST.get('phone')
            email = request.POST.get('email')
            gender = request.POST.get('gender')
            doctor_details = request.POST.get('doctor_details')
            address = request.POST.get('address')

            doctor = Doctor(
                doctor_id=doctor_id,
                doctor_name = doctor_name,
                dob = dob,
                specialization = specialization,
                experience = experience,
                age = age,
                phone = phone,
                email = email,
                gender = gender,
                doctor_details = doctor_details,
                address = address
            )
            doctor.save()
        return render(request,'doctors/add_doctor.html')
    else:
        return redirect('')


def ABOUT_DOCTOR(request):
    if request.session.has_key('id'):
        data = Doctor.objects.all()
        if request.method == "POST":
            doctor_id=request.POST.get('doctor_id')
            specialization = request.POST.get('specialization')
            experience = request.POST.get('experience')
            gender = request.POST.get('gender')
            address = request.POST.get('address')
            phone = request.POST.get('phone')
            dob = request.POST.get('dob')

            data = Doctor(doctor_id,specialization,experience,gender,address,phone,dob)
            data.save()
        return render(request,'doctors/about_doctor.html',{"details":data})
    else:
        return redirect('')


def ALL_DOCTOR(request):
    if request.session.has_key('id'):
        data = Doctor.objects.all()
        if request.method == "POST":
            doctor_id = request.POST.get('doctor_id')
            doctor_name = request.POST.get('doctor_name')
            experience = request.POST.get('experience')
            phone = request.POST.get('phone')
            specialization = request.POST.get('specialization')

            data = Doctor(doctor_id,doctor_name,experience,phone,specialization)
            data.save()
        return render(request,'doctors/all_doctor.html',{"details":data})
    else:
        return redirect('')



def EDIT_DOCTOR(request,doctor_id):
    doctor = Doctor.objects.get(doctor_id=doctor_id)

    return render(request,'doctors/edit_doctor.html',{'Doctor':doctor})

def Update_Doctor(request,doctor_id):

    doctor_name = request.POST.get('doctor_name')
    dob = request.POST.get('dob')
    specialization = request.POST.get('specialization')
    experience = request.POST.get('experience')
    age = request.POST.get('age')
    phone = request.POST.get('phone')
    email = request.POST.get('email')
    gender = request.POST.get('gender')
    doctor_details = request.POST.get('doctor_details')
    address = request.POST.get('address')
    my_user = Doctor.objects.get(doctor_id=doctor_id)
    my_user.doctor_name = doctor_name,
    my_user.dob = dob,
    my_user.specialization = specialization,
    my_user.experience = experience,
    my_user.age = age,
    my_user.phone = phone,
    my_user.email = email,
    my_user.gender = gender,
    my_user.doctor_details = doctor_details,
    my_user.address = address

    my_user.save()
    return redirect('all_doctor')


def ADD_APPOINTMENT(request):
    if request.session.has_key('id'):
        token_no = 1 if Appointment.objects.count() == 0 else Appointment.objects.aggregate(max=Max('token_no'))["max"] + 1
        if request.method == "POST":
            patient_id = request.POST.get('patient_id')
            department = request.POST.get('department')
            doctor_name = request.POST.get('doctor_name')
            appointment_data = request.POST.get('appointment_data')
            time_slot = request.POST.get('time_slot')
            appointment_id = 101 if Appointment.objects.count() == 0 else Appointment.objects.aggregate(max=Max('appointment_id'))["max"] + 1

            problem = request.POST.get('problem')

            appointment = Appointment(
                patient_id = patient_id,
                department = department,
                doctor_name = doctor_name,
                appointment_data = appointment_data,
                time_slot = time_slot,
                token_no = token_no,
                problem = problem,
                appointment_id = appointment_id,
            )
            appointment.save()
        return render(request,'appointments/add_appointments.html')
    else:
        return redirect('')


def ALL_APPOINTMENT(request):
    if request.session.has_key('id'):
        data = Appointment.objects.all()
        if request.method == "POST":
            appointment_id = request.POST.get('appointment_id')
            patient_id = request.POST.get('patient_id')
            token_no = request.POST.get('token_no')
            doctor_name = request.POST.get('doctor_name')
            problem = request.POST.get('problem')

            data = Appointment(appointment_id,patient_id,token_no,doctor_name,problem)
            data.save()
        return render(request,'appointments/all_appointments.html', {"details":data})
    else:
        return redirect('')


def ABOUT_APPOINTMENT(request):
    if request.session.has_key('id'):
        data = Appointment.objects.all()
        if request.method == "POST":
            appointment_id = request.POST.get('appointment_id')
            patient_id = request.POST.get('patient_id')
            token_no = request.POST.get('token_no')
            doctor_name = request.POST.get('doctor_name')
            problem = request.POST.get('problem')

            data = Appointment(appointment_id, patient_id, token_no, doctor_name, problem)
            data.save()
        return render(request,'appointments/details_appointments.html', {"details":data})
    else:
        return redirect('')


def EDIT_APPOINTMENT(request,appointment_id):
    appointment = Appointment.objects.get(appointment_id=appointment_id)


    return render(request,'appointments/edit_appointments.html',{'Appointment':appointment})

def Update_Appointment(request,appointment_id):
    patient_id = request.POST.get('patient_id')
    department = request.POST.get('department')
    doctor_name = request.POST.get('doctor_name')
    appointment_data = request.POST.get('appointment_data')
    time_slot = request.POST.get('time_slot')
    problem = request.POST.get('problem')
    my_user = Appointment.objects.get(appointment_id=appointment_id)
    my_user.patient_id = patient_id
    my_user.department = department
    my_user.doctor_name = doctor_name
    my_user.appointment_data = appointment_data
    my_user.time_slot = time_slot
    my_user.problem = problem

    my_user.save()
    return redirect('all_appointment')

def ADD_PAYMENT(request):
    if request.session.has_key('id'):
        if request.method == "POST":
            invoice_id = 1 if Patient_Payment.objects.count() == 0 else Patient_Payment.objects.aggregate(max=Max('invoice_id'))["max"] + 1
            invoice_date = datetime.now()
            patient_id = request.POST.get('patient_id')
            patient_name = request.POST.get('patient_name')
            department = request.POST.get('department')
            doctor_name = request.POST.get('doctor_name')
            admission_date = request.POST.get('admission_date')
            discharge_date = request.POST.get('discharge_date')
            service_name = request.POST.get('service_name')
            cost_of_treatment = request.POST.get('cost_of_treatment')
            discount = request.POST.get('discount')
            advance_paid = request.POST.get('advance_paid')
            payment_type = request.POST.get('payment_type')
            card_no = request.POST.get('card_no')

            payment = Patient_Payment(
                invoice_id = invoice_id,
                invoice_date = invoice_date,
                patient_id = patient_id,
                patient_name = patient_name,
                department = department,
                doctor_name = doctor_name,
                admission_date = admission_date,
                discharge_date = discharge_date,
                service_name = service_name,
                cost_of_treatment = cost_of_treatment,
                discount = discount,
                advance_paid = advance_paid,
                payment_type = payment_type,
                card_no = card_no,
            )
            payment.save()
        return render(request,'payment/add_payment.html')
    else:
        return redirect('')
    

def ALL_PAYMENT(request):
    if request.session.has_key('id'):
        data = Patient_Payment.objects.all()
        if request.method == "POST":
            invoice_id = request.POST.get('invoice_id')
            invoice_date = datetime.now()
            patient_name = request.POST.get('patient_name')
            doctor_name = request.POST.get('doctor_name')
            service_name = request.POST.get('service_name')
            cost_of_treatment = request.POST.get('cost_of_treatment')
            discount = request.POST.get('discount')

            data = Patient_Payment(invoice_id,invoice_date,patient_name,doctor_name,service_name,cost_of_treatment,discount)
            data.save()
        return render(request,'payment/all_payment.html', {"details":data})
    else:
        return redirect('')

def INVOICE(request):
    if request.session.has_key('id'):
        data = Patient_Payment.objects.all()
        if request.method == "POST":
            invoice_id = request.POST.get('invoice_id')
            invoice_date = datetime.now()
            patient_name = request.POST.get('patient_name')
            doctor_name = request.POST.get('doctor_name')
            service_name = request.POST.get('service_name')
            cost_of_treatment = request.POST.get('cost_of_treatment')
            discount = request.POST.get('discount')

            data = Patient_Payment(invoice_id,invoice_date,patient_name,doctor_name,service_name,cost_of_treatment,discount)
            data.save()
        return render(request,'payment/invoice.html',{'details':data})
    else:
        return redirect('')

def Invoice_E(request,invoice_id,patient_id):
    if request.session.has_key('id'):
        patient_Payment = Patient_Payment.objects.get(invoice_id=invoice_id)
        patient = Patient.objects.get(patient_id=patient_id)
        discount = patient_Payment.discount
        cost = patient_Payment.cost_of_treatment
        advance = patient_Payment.advance_paid
        dis = (discount*cost)/100
        gst = (cost*0.9)/100
        total = (gst+cost)-advance
        details= {
            'Patient_Payment':patient_Payment,
            'patient':patient,
            'total':total
        }
        print(dis)
        return render(request,'payment/invoice.html',details)
       
    else:
        return redirect('')



def invoice(request):
    invoice = Patient_Payment.objects.get()
    patient = Patient.objects.get()
    


def ADD_ROOM(request):
    if request.session.has_key('id'):
        if request.method == "POST":
            room_no = request.POST.get('room_no')
            room_type = request.POST.get('room_type')
            patient_name = request.POST.get('patient_name')
            allotment_date = request.POST.get('allotment_date')
            discharge_date = request.POST.get('discharge_date')
            doctor_name = request.POST.get('doctor_name')

            room = Room(
                room_no = room_no,
                room_type = room_type,
                patient_name = patient_name,
                allotment_date = allotment_date,
                discharge_date = discharge_date,
                doctor_name = doctor_name,
            )
            room.save()
        return render(request,'room/add_room.html')
    else:
        return redirect('')
    


def ALL_ROOM(request):
    if request.session.has_key('id'):
        data = Room.objects.all()
        if request.method == "POST":
            room_no = request.POST.get('room_no')
            room_type = request.POST.get('room_type')
            patient_name = request.POST.get('patient_name')
            allotment_date = request.POST.get('allotment_date')
            discharge_date = request.POST.get('discharge_date')
            doctor_name = request.POST.get('doctor_name')

            data = Room(room_no,room_type,patient_name,allotment_date,discharge_date,doctor_name)
            data.save()
        return render(request,'room/all_room.html', {"details":data})
    else:
        return redirect('')

def EDIT_ROOM(request,room_no):
    room = Room.objects.get(room_no = room_no)
    return render(request,'room/edit_room.html',{'Room':room})

def Update_Room(request,room_no):
    room_no = request.POST.get('room_no')
    room_type = request.POST.get('room_type')
    patient_name = request.POST.get('patient_name')
    allotment_date = request.POST.get('allotment_date')
    discharge_date = request.POST.get('discharge_date')
    doctor_name = request.POST.get('doctor_name')
    my_user = Room.objects.get(room_no=room_no)
    my_user.room_no = room_no
    my_user.room_type = room_type
    my_user.patient_name = patient_name
    my_user.allotment_date = allotment_date
    my_user.discharge_date = discharge_date
    my_user.doctor_name = doctor_name

    my_user.save()
    return redirect('all_room')


def Horizontal(request):
    if request.session.has_key('id'):
        a_data = Appointment.objects.all()
        d_data = Doctor.objects.all()
        p_total = Patient.objects.count()
        a_total = Appointment.objects.count()
        r_data = Patient_Payment.objects.aggregate(Sum('cost_of_treatment'))['cost_of_treatment__sum']
        details = {
            'a_data':a_data,
            'd_data':d_data,
            'p_total':p_total,
            'a_total':a_total,
            'r_data':r_data,
        }
        return render(request, 'dashboard/horizontal.html',details)
    else:
        return redirect('')
        


def Login(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if user:
                request.session['id'] = request.POST["username"]
                return redirect('dashboard_horizontal')
        else:
            return HttpResponse("Username or Password is incorrect!!!")
    return render(request, 'other_pages/login.html')

def SIGNUP(request):
    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
       # confirm_password = request.POST.get('confirm_password')
        my_user = User.objects.create_user(
            username, email, password
        )
        my_user.save()

        return redirect('')
    return render(request, 'other_pages/signup.html')

def Patients_chart(request):
    patients = Patients_Chart.objects.all()

    context = {
        'patients': patients

    }
    return render(request,'dashboard/horizontal.html',context)

def Logout(request):
    request.session.flush()
    return redirect('')

def p_delete(request,patient_id):
    user=Patient.objects.get(patient_id=patient_id)
    user.delete()
    return redirect('all_patient')

def d_delete(request,doctor_id):
    user=Doctor.objects.get(doctor_id=doctor_id)
    user.delete()
    return redirect('all_doctor')

def a_delete(request,appointment_id):
    user=Appointment.objects.get(appointment_id=appointment_id)
    user.delete()
    return redirect('all_appointment')

def r_delete(request,room_no):
    user=Room.objects.get(room_no=room_no)
    user.delete()
    return redirect('all_room')

class GraphData(APIView):

    def get(self,request,format=None):
        patient_chart=Patients_Chart.objects.all()
        labels = []
        DefaultData = []
        for n in patient_chart:
            labels.append(n.year)
            DefaultData.append(n.patients)
        data = {
            "labels":labels,
            "DefaultData":DefaultData,
        }
        return Response(data)
    
class GraphDataLine(APIView):

    def get(self,request,format=None):
        appointment_Chart=Appointment_Chart.objects.all()
        labels = []
        DefaultData = []
        for n in appointment_Chart:
            labels.append(n.year)
            DefaultData.append(n.appointments)
        data = {
            "labels":labels,
            "DefaultData":DefaultData,
        }
        return Response(data)


def homepage(request):
	return render(request,'dashboard/index.html')


def P_Excel(request):
    customers = Patient.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.append(['Patient Id','Patient Name','Patient Email','Patient Gender','Patient DOB','Patient Age','Phone','Address'])
    for n in customers:
        ws.append([n.patient_id,n.patient_name,n.email,n.gender,n.dob,n.age,n.phone,n.address])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_orders.xlsx'
    wb.save(response)
    return response

def D_Excel(request):
    customers = Doctor.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.append(['Doctor Id','Doctor Name','Doctor Email','Doctor Gender','Doctor DOB','Doctor Age','Specialization','Experience','Phone','Address'])
    for n in customers:
        ws.append([n.doctor_id,n.doctor_name,n.email,n.gender,n.dob,n.age,n.specialization,n.experience,n.phone,n.address])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_orders.xlsx'
    wb.save(response)
    return response

def A_Excel(request):
    customers = Appointment.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.append(['Patient Id','Appointment Id','Department','Doctor Name','Appointment Date','Time Slot','Token No','Problem'])
    for n in customers:
        ws.append([n.patient_id,n.appointment_id,n.department,n.doctor_name,n.appointment_data,n.time_slot,n.token_no,n.problem])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_orders.xlsx'
    wb.save(response)
    return response

def R_Excel(request):
    customers = Room.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.append(['Room No','Room Type','Patient Name','Allotment Date','Discharge Date','Doctor Name'])
    for n in customers:
        ws.append([n.room_no,n.room_type,n.patient_name,n.allotment_date,n.discharge_date,n.doctor_name])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_orders.xlsx'
    wb.save(response)
    return response

def PP_Excel(request):
    customers = Patient_Payment.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.append(['Invoice id','Invoice date','Patient id','Patient name','Department','Doctor name','Admission date','Discharge date','Service name','Cost of treatment','Discount','Advance paid','Payment type','Card no'])
    for n in customers:
        ws.append([n.invoice_id,n.invoice_date,n.patient_id,n.patient_name,n.department,n.doctor_name,n.admission_date,n.discharge_date,n.service_name,n.cost_of_treatment,n.discount,n.advance_paid,n.payment_type,n.card_no])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_orders.xlsx'
    wb.save(response)
    return response


    